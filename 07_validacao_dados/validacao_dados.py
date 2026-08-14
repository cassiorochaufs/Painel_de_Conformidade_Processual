#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
VALIDAÇÃO DE DADOS DO SEI 
==============================

Estratégia de verificação técnica:

1) COMPLETUDE DA AQUISIÇÃO
   SEI -> historico
   Conferência censitária das instâncias por:
   - quantidade de registros;
   - primeiro registro;
   - último registro.

2) INTEGRIDADE DAS TRANSFORMAÇÕES
   historico -> vw_cessao
   historico -> vw_cessao_fluxo

   A verificação automática é aplicada à totalidade das instâncias e compara
   propriedades observáveis entre os registros persistidos e as estruturas
   derivadas. O script não pretende certificar conformidade com normas ISO e
   não reproduz integralmente os algoritmos SQL para validar a si próprios.

3) CONTROLE MANUAL COMPLEMENTAR POR CLASSES DE COMPORTAMENTO
   O script identifica as classes efetivamente observadas e seleciona o menor
   conjunto de processos capaz de cobri-las. A conferência manual confronta
   ocorrências representativas com as estruturas derivadas aplicáveis.

Arquivos:
- Entrada manual: conferencia_sei.xlsx
- Saída consolidada: validacao_dados_saida.xlsx

A saída é sobrescrita a cada execução. Se já existir, os campos
Resultado_Manual e Observacao são reaproveitados quando a mesma classe e o
mesmo processo continuarem selecionados.

IMPORTANTE
----------
- A coluna "Descricao" não é consultada nem exportada.
- A V7 não recalcula o StayID pela mesma fórmula da vw_cessao.
- A V7 não adota percentual fixo de amostragem manual.
- Na vw_cessao_fluxo, remessas simultâneas são verificadas com multiplicidade.
- Horas_de_Transicao é validada por grupos de timestamps, pois o SQL ordena
  empates apenas por Data_Transicao.
"""

from __future__ import annotations

import math
import os
import re
from collections import Counter
from datetime import datetime
from getpass import getpass
from pathlib import Path
from typing import Any

import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import create_engine, text
from sqlalchemy.engine import URL


# =============================================================================
# CONFIGURAÇÃO
# =============================================================================

BASE_DIR = Path(__file__).resolve().parent
load_dotenv(BASE_DIR / ".env")

DB_HOST = os.getenv("DB_HOST", "localhost")
DB_PORT = int(os.getenv("DB_PORT", "5432"))
DB_NAME = os.getenv("DB_NAME", "dados_do_sei")
DB_USER = os.getenv("DB_USER", "")
DB_PASSWORD = os.getenv("DB_PASSWORD", "")

DB_SCHEMA = os.getenv("DB_SCHEMA", "public")
HIST_TABLE = os.getenv("HIST_TABLE", "historico")
VIEW_TABLE = os.getenv("VIEW_TABLE", "vw_cessao")
FLOW_VIEW_TABLE = os.getenv("FLOW_VIEW_TABLE", "vw_cessao_fluxo")

EXPECTED_PROCESS_COUNT = int(os.getenv("EXPECTED_PROCESS_COUNT", "56"))
DURATION_TOLERANCE_DAYS = float(
    os.getenv("DURATION_TOLERANCE_DAYS", "0.0001")
)
HOURS_TOLERANCE = float(os.getenv("HOURS_TOLERANCE", "0.000001"))
LOCAL_TIMEZONE = os.getenv("LOCAL_TIMEZONE", "America/Maceio")


def resolve_local_path(env_name: str, default_name: str) -> Path:
    raw = os.getenv(env_name, default_name)
    p = Path(raw)
    return p if p.is_absolute() else BASE_DIR / p


MANUAL_INPUT = resolve_local_path("MANUAL_INPUT", "conferencia_sei.xlsx")
OUTPUT_FILE = resolve_local_path("OUTPUT_FILE", "validacao_dados_saida.xlsx")


# =============================================================================
# MARCOS E CLASSES
# =============================================================================

START_TERMS = {
    "INICIO_GERADO": "processo público gerado",
    "INICIO_RECEBIDO": "processo recebido na unidade",
    "INICIO_REABERTURA": "reabertura do processo na unidade",
}

END_TERMS = {
    "FIM_CONCLUSAO": "conclusão do processo na unidade",
    "FIM_REMESSA": "processo remetido pela unidade",
    "FIM_ANEXADO": "anexado ao processo",
}

BEHAVIOR_CLASSES = [
    ("INICIO_GERADO", "Início por geração do processo"),
    ("INICIO_RECEBIDO", "Início por recebimento na unidade"),
    ("INICIO_REABERTURA", "Início por reabertura"),
    ("FIM_CONCLUSAO", "Fim por conclusão na unidade"),
    ("FIM_REMESSA", "Fim por remessa"),
    ("FIM_ANEXADO", "Fim por anexação"),
    ("RETORNO_UNIDADE", "Retorno posterior à mesma unidade"),
    ("REMESSA_SIMULTANEA", "Remessas simultâneas"),
    ("EPISODIO_EM_ANDAMENTO", "Episódio em andamento"),
]

REMESSA_RE = re.compile(
    r"remetido pela unidade\s+([A-Za-z0-9/\-]+)",
    flags=re.IGNORECASE,
)


# =============================================================================
# FUNÇÕES BÁSICAS
# =============================================================================


def validate_identifier(value: str) -> str:
    if not re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", value):
        raise ValueError(f"Identificador SQL inválido: {value!r}")
    return value


def qident(value: str) -> str:
    return f'"{validate_identifier(value)}"'


def normalize_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).casefold()


def normalize_unit(value: Any) -> str:
    return normalize_text(value)


def normalize_process(value: Any) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def timestamp_key(value: Any):
    if pd.isna(value):
        return None
    return pd.Timestamp(value).floor("min")


def parse_historico_datetime(series: pd.Series) -> pd.Series:
    def parse_one(value):
        if pd.isna(value):
            return pd.NaT

        if isinstance(value, (pd.Timestamp, datetime)):
            ts = pd.Timestamp(value)
        else:
            raw = str(value).strip()
            ts = pd.to_datetime(
                raw,
                format="%d/%m/%Y %H:%M",
                errors="coerce",
            )
            if pd.isna(ts):
                ts = pd.to_datetime(
                    raw,
                    dayfirst=True,
                    errors="coerce",
                )
            if pd.isna(ts):
                return pd.NaT
            ts = pd.Timestamp(ts)

        if ts.tzinfo is not None:
            ts = ts.tz_convert(LOCAL_TIMEZONE).tz_localize(None)

        return ts

    return series.map(parse_one)


def parse_view_datetime(series: pd.Series) -> pd.Series:
    def parse_one(value):
        if pd.isna(value):
            return pd.NaT

        ts = pd.Timestamp(value)
        if ts.tzinfo is not None:
            ts = ts.tz_convert(LOCAL_TIMEZONE).tz_localize(None)
        return ts

    return series.map(parse_one)


def parse_manual_datetime(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce", dayfirst=True)


def same_timestamp(a: Any, b: Any) -> bool | None:
    if pd.isna(a) or pd.isna(b):
        return None
    return timestamp_key(a) == timestamp_key(b)


def strip_timezone(value):
    if isinstance(value, pd.Timestamp):
        return value.tz_localize(None) if value.tzinfo is not None else value
    if isinstance(value, datetime):
        return value.replace(tzinfo=None) if value.tzinfo is not None else value
    return value


def make_excel_safe(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    for col in out.columns:
        series = out[col]

        if isinstance(series.dtype, pd.DatetimeTZDtype):
            out[col] = series.dt.tz_localize(None)
            continue

        if series.dtype == "object":
            out[col] = series.map(strip_timezone)

    return out


def counter_expanded(counter: Counter) -> list[Any]:
    rows = []
    for key, count in counter.items():
        rows.extend([key] * count)
    return rows


def floats_match_multiset(
    actual: list[float],
    expected: list[float],
    tolerance: float,
) -> bool:
    if len(actual) != len(expected):
        return False

    remaining = [float(v) for v in expected]

    for value in actual:
        found = None
        for idx, target in enumerate(remaining):
            if abs(float(value) - target) <= tolerance:
                found = idx
                break
        if found is None:
            return False
        remaining.pop(found)

    return not remaining


# =============================================================================
# CLASSIFICAÇÃO INDIVIDUAL DOS REGISTROS DO HISTÓRICO
# =============================================================================


def classify_event(obs: Any, unidade: Any) -> dict[str, Any]:
    """
    Classifica somente o registro individual.

    Não calcula StayID.
    Não reconstrói episódios.
    """
    obs_raw = "" if pd.isna(obs) else str(obs).strip()
    obs_norm = normalize_text(obs_raw)
    unidade_raw = "" if pd.isna(unidade) else str(unidade).strip()

    classe = "OUTRO"
    tipo = "OUTRO"
    unidade_evento = unidade_raw
    is_remessa = False

    for class_code, term in START_TERMS.items():
        if term in obs_norm:
            classe = class_code
            tipo = "INICIO"
            break

    if tipo == "OUTRO":
        for class_code, term in END_TERMS.items():
            if term in obs_norm:
                classe = class_code
                tipo = "FIM"
                break

    if classe == "FIM_REMESSA":
        is_remessa = True
        match = REMESSA_RE.search(obs_raw)
        if match:
            unidade_evento = match.group(1).strip()
        else:
            unidade_evento = unidade_raw

    return {
        "_Classe_Evento": classe,
        "_Tipo_Evento": tipo,
        "_Unidade_Evento": unidade_evento,
        "_Eh_Remessa": is_remessa,
    }


# =============================================================================
# BANCO
# =============================================================================


def get_engine():
    global DB_PASSWORD

    if not DB_USER:
        raise RuntimeError(
            "Defina DB_USER no arquivo .env localizado na mesma pasta do script."
        )

    if not DB_PASSWORD:
        DB_PASSWORD = getpass("Senha do PostgreSQL: ")

    url = URL.create(
        "postgresql+psycopg2",
        username=DB_USER,
        password=DB_PASSWORD,
        host=DB_HOST,
        port=DB_PORT,
        database=DB_NAME,
    )

    return create_engine(url, future=True)


def load_data(engine) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    schema = qident(DB_SCHEMA)
    historico = qident(HIST_TABLE)
    vw = qident(VIEW_TABLE)
    fluxo = qident(FLOW_VIEW_TABLE)

    sql_view = text(
        f"""
        SELECT
            "Processo",
            "Unidade",
            "StayID",
            "Data_Inicio",
            "Data_Fim",
            "Status_Etapa",
            "Duracao_Decimal_Dias"
        FROM {schema}.{vw}
        """
    )

    sql_flow = text(
        f"""
        SELECT
            "Processo",
            "Unidade_Origem",
            "Unidade_Destino",
            "Data_Transicao",
            "Horas_de_Transicao"
        FROM {schema}.{fluxo}
        """
    )

    sql_hist = text(
        f"""
        SELECT
            h."Processo",
            h."Marcador",
            h."Data",
            h."Unidade",
            h."Obs"
        FROM {schema}.{historico} AS h
        WHERE h."Processo" IN (
            SELECT DISTINCT h2."Processo"
            FROM {schema}.{historico} AS h2
            WHERE TRIM(h2."Marcador") ~* '^Marcador\\s+CESSÃO$'
        )
        """
    )

    return (
        pd.read_sql(sql_hist, engine),
        pd.read_sql(sql_view, engine),
        pd.read_sql(sql_flow, engine),
    )


# =============================================================================
# PREPARAÇÃO
# =============================================================================


def prepare_data(
    hist: pd.DataFrame,
    vw: pd.DataFrame,
    fluxo: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:

    hist = hist.copy()
    vw = vw.copy()
    fluxo = fluxo.copy()

    hist["Processo"] = hist["Processo"].map(normalize_process)
    vw["Processo"] = vw["Processo"].map(normalize_process)
    fluxo["Processo"] = fluxo["Processo"].map(normalize_process)

    hist["Data"] = parse_historico_datetime(hist["Data"])
    vw["Data_Inicio"] = parse_view_datetime(vw["Data_Inicio"])
    vw["Data_Fim"] = parse_view_datetime(vw["Data_Fim"])
    fluxo["Data_Transicao"] = parse_view_datetime(fluxo["Data_Transicao"])

    vw["StayID"] = pd.to_numeric(vw["StayID"], errors="coerce").astype("Int64")
    vw["Duracao_Decimal_Dias"] = pd.to_numeric(
        vw["Duracao_Decimal_Dias"],
        errors="coerce",
    )
    fluxo["Horas_de_Transicao"] = pd.to_numeric(
        fluxo["Horas_de_Transicao"],
        errors="coerce",
    )

    classified = hist.apply(
        lambda row: pd.Series(
            classify_event(row["Obs"], row["Unidade"])
        ),
        axis=1,
    )

    hist = pd.concat([hist, classified], axis=1)

    hist["_Unidade_Evento_Norm"] = hist["_Unidade_Evento"].map(normalize_unit)
    hist["_Unidade_Norm"] = hist["Unidade"].map(normalize_unit)

    vw["_Unidade_Norm"] = vw["Unidade"].map(normalize_unit)
    vw["_Status_Norm"] = vw["Status_Etapa"].map(normalize_text)

    fluxo["_Origem_Norm"] = fluxo["Unidade_Origem"].map(normalize_unit)
    fluxo["_Destino_Norm"] = fluxo["Unidade_Destino"].map(normalize_unit)

    return hist, vw, fluxo


# =============================================================================
# 1. COMPLETUDE DA AQUISIÇÃO
# =============================================================================


def ensure_manual_input(processes: list[str]) -> bool:
    if MANUAL_INPUT.exists():
        return False

    template = pd.DataFrame(
        {
            "Processo": processes,
            "Qtd_Registros_SEI": pd.Series(
                [pd.NA] * len(processes),
                dtype="Int64",
            ),
            "Primeiro_Registro_SEI": [pd.NaT] * len(processes),
            "Ultimo_Registro_SEI": [pd.NaT] * len(processes),
            "Data_Conferencia_SEI": [pd.NaT] * len(processes),
            "Observacao": [""] * len(processes),
        }
    )

    with pd.ExcelWriter(
        MANUAL_INPUT,
        engine="openpyxl",
        datetime_format="dd/mm/yyyy hh:mm",
    ) as writer:
        template.to_excel(
            writer,
            index=False,
            sheet_name="CONFERENCIA_SEI",
        )

    style_workbook(MANUAL_INPUT)
    return True


def build_completeness(
    hist: pd.DataFrame,
    processes: list[str],
) -> pd.DataFrame:

    hist_summary = (
        hist.groupby("Processo", dropna=False)
        .agg(
            Qtd_Registros_Historico=("Processo", "size"),
            Primeiro_Registro_Historico=("Data", "min"),
            Ultimo_Registro_Historico=("Data", "max"),
        )
        .reset_index()
    )

    manual = pd.read_excel(
        MANUAL_INPUT,
        sheet_name="CONFERENCIA_SEI",
    )

    manual["Processo"] = manual["Processo"].map(normalize_process)
    manual["Qtd_Registros_SEI"] = pd.to_numeric(
        manual["Qtd_Registros_SEI"],
        errors="coerce",
    )
    manual["Primeiro_Registro_SEI"] = parse_manual_datetime(
        manual["Primeiro_Registro_SEI"]
    )
    manual["Ultimo_Registro_SEI"] = parse_manual_datetime(
        manual["Ultimo_Registro_SEI"]
    )

    out = (
        pd.DataFrame({"Processo": processes})
        .merge(manual, on="Processo", how="left")
        .merge(hist_summary, on="Processo", how="left")
    )

    out["Diferenca_Registros"] = (
        out["Qtd_Registros_Historico"]
        - out["Qtd_Registros_SEI"]
    )

    qtd_confere = []
    first_confere = []
    last_confere = []
    final = []

    for row in out.itertuples(index=False):
        qtd_ok = (
            None
            if pd.isna(row.Qtd_Registros_SEI)
            else row.Qtd_Registros_SEI == row.Qtd_Registros_Historico
        )

        first_ok = same_timestamp(
            row.Primeiro_Registro_SEI,
            row.Primeiro_Registro_Historico,
        )

        last_ok = same_timestamp(
            row.Ultimo_Registro_SEI,
            row.Ultimo_Registro_Historico,
        )

        def flag(v):
            if v is None:
                return "PENDENTE"
            return "SIM" if v else "NÃO"

        qtd_confere.append(flag(qtd_ok))
        first_confere.append(flag(first_ok))
        last_confere.append(flag(last_ok))

        checks = [qtd_ok, first_ok, last_ok]

        if any(v is None for v in checks):
            final.append("PENDENTE")
        elif all(checks):
            final.append("OK")
        else:
            final.append("CONFERIR")

    out["Qtd_Confere"] = qtd_confere
    out["Primeiro_Registro_Confere"] = first_confere
    out["Ultimo_Registro_Confere"] = last_confere
    out["Resultado_Completude"] = final

    return out[
        [
            "Processo",
            "Qtd_Registros_SEI",
            "Qtd_Registros_Historico",
            "Diferenca_Registros",
            "Primeiro_Registro_SEI",
            "Primeiro_Registro_Historico",
            "Primeiro_Registro_Confere",
            "Ultimo_Registro_SEI",
            "Ultimo_Registro_Historico",
            "Ultimo_Registro_Confere",
            "Qtd_Confere",
            "Resultado_Completude",
            "Data_Conferencia_SEI",
            "Observacao",
        ]
    ]


# =============================================================================
# 2A. VERIFICAÇÃO AUTOMÁTICA DA vw_cessao
# =============================================================================


def run_automatic_validation_cessao(
    hist: pd.DataFrame,
    vw: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:

    hist_processes = set(hist["Processo"].dropna().unique())
    view_processes = set(vw["Processo"].dropna().unique())
    processes = sorted(hist_processes | view_processes)

    rules = ["V01", "V02", "V03", "V04", "V05", "V06", "V07"]

    result = pd.DataFrame({"Processo": processes})
    for rule in rules:
        result[rule] = "OK"

    divergences: list[dict[str, Any]] = []

    def fail(
        proc: str,
        rule: str,
        detail: str,
        unit: str = "",
        dt: Any = None,
    ):
        if proc in set(result["Processo"]):
            result.loc[result["Processo"] == proc, rule] = "FALHA"

        divergences.append(
            {
                "Estrutura": "vw_cessao",
                "Processo": proc,
                "Regra": rule,
                "Unidade_Origem": unit,
                "Unidade_Destino": "",
                "Data": pd.Timestamp(dt) if pd.notna(dt) else pd.NaT,
                "Detalhe": detail,
            }
        )

    # V01 - cobertura do universo
    for proc in sorted(hist_processes - view_processes):
        fail(
            proc,
            "V01",
            "Processo de cessão presente em historico e ausente na vw_cessao.",
        )

    for proc in sorted(view_processes - hist_processes):
        fail(
            proc,
            "V01",
            "Processo presente na vw_cessao e ausente do universo de cessão em historico.",
        )

    start_events = hist[hist["_Tipo_Evento"] == "INICIO"].copy()
    end_events = hist[hist["_Tipo_Evento"] == "FIM"].copy()
    remessa_events = hist[hist["_Eh_Remessa"]].copy()

    start_keys = {
        (
            proc,
            normalize_unit(unit),
            timestamp_key(dt),
        )
        for proc, unit, dt in start_events[
            ["Processo", "_Unidade_Evento", "Data"]
        ].itertuples(index=False, name=None)
        if pd.notna(dt)
    }

    end_keys = {
        (
            proc,
            normalize_unit(unit),
            timestamp_key(dt),
        )
        for proc, unit, dt in end_events[
            ["Processo", "_Unidade_Evento", "Data"]
        ].itertuples(index=False, name=None)
        if pd.notna(dt)
    }

    remessa_keys = {
        (
            proc,
            normalize_unit(unit),
            timestamp_key(dt),
        )
        for proc, unit, dt in remessa_events[
            ["Processo", "_Unidade_Evento", "Data"]
        ].itertuples(index=False, name=None)
        if pd.notna(dt)
    }

    # V02 e V03
    for proc, unit, stay, start, end, status, duration in vw[
        [
            "Processo",
            "Unidade",
            "StayID",
            "Data_Inicio",
            "Data_Fim",
            "Status_Etapa",
            "Duracao_Decimal_Dias",
        ]
    ].itertuples(index=False, name=None):

        start_key = (
            proc,
            normalize_unit(unit),
            timestamp_key(start),
        )

        if start_key not in start_keys:
            fail(
                proc,
                "V02",
                "Data_Inicio sem marco de INÍCIO correspondente em historico.",
                str(unit),
                start,
            )

        if pd.notna(end):
            end_key = (
                proc,
                normalize_unit(unit),
                timestamp_key(end),
            )

            if end_key not in end_keys:
                fail(
                    proc,
                    "V03",
                    "Data_Fim sem marco de FIM correspondente em historico.",
                    str(unit),
                    end,
                )

    # V04 - tratamento das remessas
    # V03 já confronta a unidade do episódio com a origem extraída de Obs.
    # V04 verifica a pré-condição dessa regra: a unidade remetente precisa ser
    # identificável no texto de cada evento de remessa. Isso evita que um
    # fallback silencioso para a unidade de destino seja interpretado como
    # evidência suficiente de rastreabilidade.
    for proc, obs, unidade, dt in hist.loc[
        hist["_Eh_Remessa"],
        ["Processo", "Obs", "Unidade", "Data"],
    ].itertuples(index=False, name=None):
        obs_raw = "" if pd.isna(obs) else str(obs)
        if REMESSA_RE.search(obs_raw) is None:
            fail(
                proc,
                "V04",
                "Evento de remessa sem unidade remetente identificável no campo Obs.",
                "" if pd.isna(unidade) else str(unidade),
                dt,
            )

    # V05 - coerência temporal e de status
    for proc, unit, start, end, status in vw[
        [
            "Processo",
            "Unidade",
            "Data_Inicio",
            "Data_Fim",
            "Status_Etapa",
        ]
    ].itertuples(index=False, name=None):

        status_norm = normalize_text(status)

        if pd.notna(start) and pd.notna(end) and end < start:
            fail(
                proc,
                "V05",
                "Data_Fim anterior à Data_Inicio.",
                str(unit),
                end,
            )

        if status_norm == "concluída" and pd.isna(end):
            fail(
                proc,
                "V05",
                'Status "Concluída" com Data_Fim vazia.',
                str(unit),
            )

        elif status_norm == "em andamento" and pd.notna(end):
            fail(
                proc,
                "V05",
                'Status "Em Andamento" com Data_Fim preenchida.',
                str(unit),
                end,
            )

        elif status_norm not in {"concluída", "em andamento"}:
            fail(
                proc,
                "V05",
                f"Status_Etapa inesperado: {status!r}.",
                str(unit),
            )

    # V06 - duração
    concluded = vw[vw["_Status_Norm"] == "concluída"]

    for proc, unit, start, end, duration in concluded[
        [
            "Processo",
            "Unidade",
            "Data_Inicio",
            "Data_Fim",
            "Duracao_Decimal_Dias",
        ]
    ].itertuples(index=False, name=None):

        if pd.isna(start) or pd.isna(end):
            continue

        calculated = (end - start).total_seconds() / 86400.0

        if pd.isna(duration):
            fail(
                proc,
                "V06",
                "Episódio concluído sem Duracao_Decimal_Dias.",
                str(unit),
                end,
            )

        elif abs(float(duration) - calculated) > DURATION_TOLERANCE_DAYS:
            fail(
                proc,
                "V06",
                (
                    f"Duração divergente. Calculada={calculated:.6f}; "
                    f"vw_cessao={float(duration):.6f}."
                ),
                str(unit),
                end,
            )

    # V07 - integridade dos episódios
    essential_missing = vw[
        vw["Processo"].eq("")
        | vw["Unidade"].isna()
        | vw["StayID"].isna()
        | vw["Data_Inicio"].isna()
    ]

    for proc, unit, stay, start in essential_missing[
        ["Processo", "Unidade", "StayID", "Data_Inicio"]
    ].itertuples(index=False, name=None):

        fail(
            proc if proc else "SEM_PROCESSO",
            "V07",
            "Campo essencial ausente: Processo, Unidade, StayID ou Data_Inicio.",
            "" if pd.isna(unit) else str(unit),
            start,
        )

    duplicate_cols = [
        "Processo",
        "_Unidade_Norm",
        "StayID",
        "Data_Inicio",
        "Data_Fim",
    ]

    duplicates = vw[
        vw.duplicated(
            subset=duplicate_cols,
            keep=False,
        )
    ]

    for proc, unit, stay, start in duplicates[
        ["Processo", "Unidade", "StayID", "Data_Inicio"]
    ].itertuples(index=False, name=None):

        fail(
            proc,
            "V07",
            "Duplicação exata de episódio identificada.",
            str(unit),
            start,
        )

    hist_counts = hist.groupby("Processo").size()
    view_counts = vw.groupby("Processo").size()

    result["Qtd_Registros_Historico"] = (
        result["Processo"].map(hist_counts).fillna(0).astype(int)
    )
    result["Qtd_Episodios_vw_cessao"] = (
        result["Processo"].map(view_counts).fillna(0).astype(int)
    )

    result["Qtd_Falhas"] = (
        result[rules] == "FALHA"
    ).sum(axis=1)

    result["Resultado_Automatico"] = result["Qtd_Falhas"].apply(
        lambda n: "OK" if n == 0 else "CONFERIR"
    )

    result = result[
        [
            "Processo",
            "Qtd_Registros_Historico",
            "Qtd_Episodios_vw_cessao",
            *rules,
            "Qtd_Falhas",
            "Resultado_Automatico",
        ]
    ]

    divergence_df = pd.DataFrame(
        divergences,
        columns=[
            "Estrutura",
            "Processo",
            "Regra",
            "Unidade_Origem",
            "Unidade_Destino",
            "Data",
            "Detalhe",
        ],
    )

    return result, divergence_df


# =============================================================================
# 2B. VERIFICAÇÃO AUTOMÁTICA DA vw_cessao_fluxo
# =============================================================================


def run_automatic_validation_fluxo(
    hist: pd.DataFrame,
    fluxo: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Verifica a vw_cessao_fluxo contra propriedades observáveis do historico.

    O procedimento não replica ROW_NUMBER + CASE + LAG como uma segunda
    implementação completa da view. Em vez disso, verifica cobertura,
    transição inicial, rastreabilidade/multiplicidade das remessas e fins,
    ausência de origem desconhecida, coerência temporal dos intervalos e
    integridade estrutural.
    """

    hist_processes = set(hist["Processo"].dropna().unique())
    flow_processes = set(fluxo["Processo"].dropna().unique())
    processes = sorted(hist_processes | flow_processes)

    rules = ["F01", "F02", "F03", "F04", "F05", "F06", "F07"]

    result = pd.DataFrame({"Processo": processes})
    for rule in rules:
        result[rule] = "OK"

    divergences: list[dict[str, Any]] = []

    def fail(
        proc: str,
        rule: str,
        detail: str,
        origin: str = "",
        destination: str = "",
        dt: Any = None,
    ):
        if proc in set(result["Processo"]):
            result.loc[result["Processo"] == proc, rule] = "FALHA"

        divergences.append(
            {
                "Estrutura": "vw_cessao_fluxo",
                "Processo": proc,
                "Regra": rule,
                "Unidade_Origem": origin,
                "Unidade_Destino": destination,
                "Data": pd.Timestamp(dt) if pd.notna(dt) else pd.NaT,
                "Detalhe": detail,
            }
        )

    # F01 - cobertura do universo
    for proc in sorted(hist_processes - flow_processes):
        fail(
            proc,
            "F01",
            "Processo de cessão presente em historico e ausente na vw_cessao_fluxo.",
        )

    for proc in sorted(flow_processes - hist_processes):
        fail(
            proc,
            "F01",
            "Processo presente na vw_cessao_fluxo e ausente do universo de cessão em historico.",
        )

    # F02 - transição inicial
    initial_map: dict[str, tuple[Any, str]] = {}

    for proc in processes:
        h = hist[hist["Processo"] == proc]
        f = fluxo[fluxo["Processo"] == proc]

        if h.empty:
            continue

        min_dt = h["Data"].min()
        first_candidates = h[h["Data"] == min_dt]
        candidate_units = set(first_candidates["_Unidade_Norm"].tolist())

        initials = f[f["_Origem_Norm"] == "início"]

        if len(initials) != 1:
            fail(
                proc,
                "F02",
                f"Esperada exatamente 1 transição inicial; encontradas {len(initials)}.",
            )
            continue

        row = initials.iloc[0]
        init_dt = row["Data_Transicao"]
        init_dest = row["_Destino_Norm"]
        initial_map[proc] = (timestamp_key(init_dt), init_dest)

        if timestamp_key(init_dt) != timestamp_key(min_dt):
            fail(
                proc,
                "F02",
                "A transição INÍCIO não ocorre no primeiro instante registrado no historico.",
                str(row["Unidade_Origem"]),
                str(row["Unidade_Destino"]),
                init_dt,
            )

        if init_dest not in candidate_units:
            fail(
                proc,
                "F02",
                "O destino da transição INÍCIO não corresponde a uma unidade registrada no primeiro instante do processo.",
                str(row["Unidade_Origem"]),
                str(row["Unidade_Destino"]),
                init_dt,
            )

    # Construção de contadores observáveis para F03/F04.
    # Remessas: origem extraída de Obs, destino = Unidade do historico.
    raw_remessa_by_proc: dict[str, Counter] = {}
    raw_end_by_proc: dict[str, Counter] = {}
    flow_remessa_by_proc: dict[str, Counter] = {}
    flow_end_by_proc: dict[str, Counter] = {}

    for proc in processes:
        raw_remessa_by_proc[proc] = Counter()
        raw_end_by_proc[proc] = Counter()
        flow_remessa_by_proc[proc] = Counter()
        flow_end_by_proc[proc] = Counter()

    for proc, origem, destino, dt in hist.loc[
        hist["_Eh_Remessa"],
        ["Processo", "_Unidade_Evento", "Unidade", "Data"],
    ].itertuples(index=False, name=None):
        key = (
            normalize_unit(origem),
            normalize_unit(destino),
            timestamp_key(dt),
        )
        raw_remessa_by_proc.setdefault(proc, Counter())[key] += 1

    end_mask = hist["_Classe_Evento"].isin({"FIM_CONCLUSAO", "FIM_ANEXADO"})
    for proc, unidade, dt in hist.loc[
        end_mask,
        ["Processo", "Unidade", "Data"],
    ].itertuples(index=False, name=None):
        key = (
            normalize_unit(unidade),
            "fim",
            timestamp_key(dt),
        )
        raw_end_by_proc.setdefault(proc, Counter())[key] += 1

    for proc, origin, dest, dt in fluxo[
        ["Processo", "Unidade_Origem", "Unidade_Destino", "Data_Transicao"]
    ].itertuples(index=False, name=None):
        origin_n = normalize_unit(origin)
        dest_n = normalize_unit(dest)
        key = (origin_n, dest_n, timestamp_key(dt))

        if origin_n == "início":
            continue
        if dest_n == "fim":
            flow_end_by_proc.setdefault(proc, Counter())[key] += 1
        else:
            flow_remessa_by_proc.setdefault(proc, Counter())[key] += 1

    # F03/F04 - comparação com multiplicidade e exceção da primeira linha.
    # Se o primeiro registro do processo também for uma remessa/conclusão/anexação,
    # o CASE da view dá precedência a INÍCIO. Quando há empate de timestamp no
    # primeiro instante, o ROW_NUMBER do SQL não define desempate adicional.
    # Por isso, admite-se no máximo uma ocorrência elegível não representada como
    # remessa/fim, desde que ela seja compatível com a transição inicial observada.
    for proc in processes:
        raw_r = raw_remessa_by_proc.get(proc, Counter())
        raw_e = raw_end_by_proc.get(proc, Counter())
        flow_r = flow_remessa_by_proc.get(proc, Counter())
        flow_e = flow_end_by_proc.get(proc, Counter())

        missing_r = raw_r - flow_r
        extra_r = flow_r - raw_r
        missing_e = raw_e - flow_e
        extra_e = flow_e - raw_e

        missing_r_list = counter_expanded(missing_r)
        missing_e_list = counter_expanded(missing_e)

        # Aplicar uma única tolerância de precedência da primeira linha.
        init = initial_map.get(proc)
        if init and len(missing_r_list) + len(missing_e_list) == 1:
            init_dt, init_dest = init

            if missing_r_list:
                origin_n, dest_n, dt = missing_r_list[0]
                if dt == init_dt and dest_n == init_dest:
                    missing_r = Counter()
                    missing_r_list = []

            elif missing_e_list:
                origin_n, dest_n, dt = missing_e_list[0]
                if dt == init_dt and origin_n == init_dest:
                    missing_e = Counter()
                    missing_e_list = []

        if missing_r or extra_r:
            fail(
                proc,
                "F03",
                (
                    "Divergência nas remessas origem-destino. "
                    f"Ausentes na view={sum(missing_r.values())}; "
                    f"excedentes na view={sum(extra_r.values())}."
                ),
            )

        if missing_e or extra_e:
            fail(
                proc,
                "F04",
                (
                    "Divergência nas transições de conclusão/anexação para FIM. "
                    f"Ausentes na view={sum(missing_e.values())}; "
                    f"excedentes na view={sum(extra_e.values())}."
                ),
            )

    # F05 - origem desconhecida em remessa
    unknown = fluxo[fluxo["_Origem_Norm"] == "desconhecida"]
    for proc, origin, dest, dt in unknown[
        ["Processo", "Unidade_Origem", "Unidade_Destino", "Data_Transicao"]
    ].itertuples(index=False, name=None):
        fail(
            proc,
            "F05",
            "Origem DESCONHECIDA identificada na vw_cessao_fluxo.",
            str(origin),
            str(dest),
            dt,
        )

    # F06 - Horas_de_Transicao por grupos de timestamps.
    # Em empates, ORDER BY Data_Transicao não define a ordem interna. Portanto,
    # o teste compara o multiconjunto esperado dentro de cada timestamp:
    # - primeiro timestamp: um NULL e zeros para as demais linhas simultâneas;
    # - demais timestamps: um delta desde o timestamp anterior e zeros para as
    #   demais linhas simultâneas.
    for proc in processes:
        f = fluxo[fluxo["Processo"] == proc].copy()
        if f.empty:
            continue

        if f["Data_Transicao"].isna().any():
            continue

        distinct_times = sorted(f["Data_Transicao"].dropna().unique())
        previous_time = None

        for idx, current_time in enumerate(distinct_times):
            group = f[f["Data_Transicao"] == current_time]
            hours = group["Horas_de_Transicao"]
            null_count = int(hours.isna().sum())
            actual_nonnull = [float(v) for v in hours.dropna().tolist()]

            if idx == 0:
                expected_null = 1
                expected_nonnull = [0.0] * (len(group) - 1)
            else:
                expected_null = 0
                delta = (
                    pd.Timestamp(current_time) - pd.Timestamp(previous_time)
                ).total_seconds() / 3600.0
                expected_nonnull = [float(delta)] + [0.0] * (len(group) - 1)

            if (
                null_count != expected_null
                or not floats_match_multiset(
                    actual_nonnull,
                    expected_nonnull,
                    HOURS_TOLERANCE,
                )
            ):
                fail(
                    proc,
                    "F06",
                    (
                        "Horas_de_Transicao incoerente no grupo temporal. "
                        f"Timestamp={pd.Timestamp(current_time)}; "
                        f"valores={hours.tolist()}."
                    ),
                    dt=current_time,
                )

            previous_time = current_time

    # F07 - integridade estrutural
    essential_missing = fluxo[
        fluxo["Processo"].eq("")
        | fluxo["Unidade_Origem"].isna()
        | fluxo["Unidade_Destino"].isna()
        | fluxo["Data_Transicao"].isna()
        | fluxo["_Origem_Norm"].eq("")
        | fluxo["_Destino_Norm"].eq("")
    ]

    for proc, origin, dest, dt in essential_missing[
        ["Processo", "Unidade_Origem", "Unidade_Destino", "Data_Transicao"]
    ].itertuples(index=False, name=None):
        fail(
            proc if proc else "SEM_PROCESSO",
            "F07",
            "Campo essencial ausente em transição da vw_cessao_fluxo.",
            "" if pd.isna(origin) else str(origin),
            "" if pd.isna(dest) else str(dest),
            dt,
        )

    negative_hours = fluxo[
        fluxo["Horas_de_Transicao"].notna()
        & (fluxo["Horas_de_Transicao"] < -HOURS_TOLERANCE)
    ]

    for proc, origin, dest, dt, hours in negative_hours[
        [
            "Processo",
            "Unidade_Origem",
            "Unidade_Destino",
            "Data_Transicao",
            "Horas_de_Transicao",
        ]
    ].itertuples(index=False, name=None):
        fail(
            proc,
            "F07",
            f"Horas_de_Transicao negativa: {hours}.",
            str(origin),
            str(dest),
            dt,
        )

    hist_counts = hist.groupby("Processo").size()
    flow_counts = fluxo.groupby("Processo").size()

    result["Qtd_Registros_Historico"] = (
        result["Processo"].map(hist_counts).fillna(0).astype(int)
    )
    result["Qtd_Transicoes_vw_fluxo"] = (
        result["Processo"].map(flow_counts).fillna(0).astype(int)
    )

    result["Qtd_Falhas"] = (
        result[rules] == "FALHA"
    ).sum(axis=1)

    result["Resultado_Automatico"] = result["Qtd_Falhas"].apply(
        lambda n: "OK" if n == 0 else "CONFERIR"
    )

    result = result[
        [
            "Processo",
            "Qtd_Registros_Historico",
            "Qtd_Transicoes_vw_fluxo",
            *rules,
            "Qtd_Falhas",
            "Resultado_Automatico",
        ]
    ]

    divergence_df = pd.DataFrame(
        divergences,
        columns=[
            "Estrutura",
            "Processo",
            "Regra",
            "Unidade_Origem",
            "Unidade_Destino",
            "Data",
            "Detalhe",
        ],
    )

    return result, divergence_df


# =============================================================================
# 3. CLASSES DE COMPORTAMENTO
# =============================================================================


def build_behavior_occurrences(
    hist: pd.DataFrame,
    vw: pd.DataFrame,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []

    # Classes diretamente observáveis em historico
    for class_code, class_name in BEHAVIOR_CLASSES[:6]:
        subset = hist[hist["_Classe_Evento"] == class_code]

        for proc, unit_evento, dt, obs in subset[
            ["Processo", "_Unidade_Evento", "Data", "Obs"]
        ].itertuples(index=False, name=None):

            rows.append(
                {
                    "Processo": proc,
                    "Classe": class_code,
                    "Classe_Descricao": class_name,
                    "Unidade": unit_evento,
                    "Data": dt,
                    "Evidencia": str(obs),
                }
            )

    # Retorno à mesma unidade
    repeated = (
        vw.groupby(["Processo", "_Unidade_Norm"], dropna=False)
        .size()
        .reset_index(name="Qtd_Episodios")
    )

    repeated = repeated[repeated["Qtd_Episodios"] > 1]

    return_name = dict(BEHAVIOR_CLASSES)["RETORNO_UNIDADE"]

    for proc, unit_norm, qtd in repeated[
        ["Processo", "_Unidade_Norm", "Qtd_Episodios"]
    ].itertuples(index=False, name=None):

        group = vw[
            (vw["Processo"] == proc)
            & (vw["_Unidade_Norm"] == unit_norm)
        ].sort_values("Data_Inicio")

        first_date = (
            group["Data_Inicio"].iloc[0]
            if not group.empty
            else pd.NaT
        )

        stay_ids = ", ".join(
            str(int(v))
            for v in group["StayID"].dropna().tolist()
        )

        rows.append(
            {
                "Processo": proc,
                "Classe": "RETORNO_UNIDADE",
                "Classe_Descricao": return_name,
                "Unidade": (
                    str(group["Unidade"].iloc[0])
                    if not group.empty
                    else unit_norm
                ),
                "Data": first_date,
                "Evidencia": (
                    f"{qtd} episódios na mesma unidade; StayID: {stay_ids}"
                ),
            }
        )

    # Remessas simultâneas
    remessas = hist[hist["_Eh_Remessa"]].copy()

    simultaneous_name = dict(BEHAVIOR_CLASSES)["REMESSA_SIMULTANEA"]

    if not remessas.empty:
        simultaneous = (
            remessas.groupby(["Processo", "Data"], dropna=False)
            .size()
            .reset_index(name="Qtd_Remessas")
        )

        simultaneous = simultaneous[
            simultaneous["Qtd_Remessas"] > 1
        ]

        for proc, dt, qtd in simultaneous[
            ["Processo", "Data", "Qtd_Remessas"]
        ].itertuples(index=False, name=None):

            origins = remessas[
                (remessas["Processo"] == proc)
                & (remessas["Data"] == dt)
            ]["_Unidade_Evento"].astype(str).tolist()

            destinations = remessas[
                (remessas["Processo"] == proc)
                & (remessas["Data"] == dt)
            ]["Unidade"].astype(str).tolist()

            pairs = [
                f"{o} -> {d}"
                for o, d in zip(origins, destinations)
            ]

            rows.append(
                {
                    "Processo": proc,
                    "Classe": "REMESSA_SIMULTANEA",
                    "Classe_Descricao": simultaneous_name,
                    "Unidade": "; ".join(pairs),
                    "Data": dt,
                    "Evidencia": (
                        f"{qtd} remessas no mesmo timestamp: "
                        + "; ".join(pairs)
                    ),
                }
            )

    # Episódio em andamento
    open_name = dict(BEHAVIOR_CLASSES)["EPISODIO_EM_ANDAMENTO"]
    open_rows = vw[vw["_Status_Norm"] == "em andamento"]

    for proc, unit, start, stay in open_rows[
        ["Processo", "Unidade", "Data_Inicio", "StayID"]
    ].itertuples(index=False, name=None):

        rows.append(
            {
                "Processo": proc,
                "Classe": "EPISODIO_EM_ANDAMENTO",
                "Classe_Descricao": open_name,
                "Unidade": unit,
                "Data": start,
                "Evidencia": (
                    f"StayID={stay}; Status_Etapa=Em Andamento."
                ),
            }
        )

    return pd.DataFrame(
        rows,
        columns=[
            "Processo",
            "Classe",
            "Classe_Descricao",
            "Unidade",
            "Data",
            "Evidencia",
        ],
    )


def build_class_summary(
    occurrences: pd.DataFrame,
) -> pd.DataFrame:

    rows = []

    for class_code, class_name in BEHAVIOR_CLASSES:
        subset = occurrences[
            occurrences["Classe"] == class_code
        ]

        rows.append(
            {
                "Classe": class_code,
                "Classe_Descricao": class_name,
                "Ocorrencias": len(subset),
                "Processos": subset["Processo"].nunique(),
                "Status": (
                    "OBSERVADA"
                    if not subset.empty
                    else "NÃO OBSERVADA NO RECORTE"
                ),
            }
        )

    return pd.DataFrame(rows)


def select_minimum_process_cover(
    occurrences: pd.DataFrame,
) -> list[str]:
    """
    Seleciona exatamente o menor número de processos necessário para cobrir
    todas as classes de comportamento observadas.
    """
    observed_classes = sorted(
        occurrences["Classe"].dropna().unique().tolist()
    )

    if not observed_classes:
        return []

    class_to_bit = {
        class_code: idx
        for idx, class_code in enumerate(observed_classes)
    }

    process_classes = (
        occurrences.groupby("Processo")["Classe"]
        .apply(lambda s: set(s.tolist()))
        .to_dict()
    )

    process_masks = {}

    for proc, classes in process_classes.items():
        mask = 0
        for class_code in classes:
            if class_code in class_to_bit:
                mask |= 1 << class_to_bit[class_code]
        process_masks[proc] = mask

    target = (1 << len(observed_classes)) - 1

    dp: dict[int, tuple[str, ...]] = {0: tuple()}

    for proc in sorted(process_masks):
        p_mask = process_masks[proc]
        snapshot = list(dp.items())

        for mask, chosen in snapshot:
            new_mask = mask | p_mask
            candidate = chosen + (proc,)

            current = dp.get(new_mask)

            if (
                current is None
                or len(candidate) < len(current)
                or (
                    len(candidate) == len(current)
                    and candidate < current
                )
            ):
                dp[new_mask] = candidate

    return list(dp.get(target, tuple()))


def describe_structures_for_class(class_code: str) -> str:
    mapping = {
        "INICIO_GERADO": "historico ↔ vw_cessao",
        "INICIO_RECEBIDO": "historico ↔ vw_cessao",
        "INICIO_REABERTURA": "historico ↔ vw_cessao",
        "FIM_CONCLUSAO": "historico ↔ vw_cessao ↔ vw_cessao_fluxo",
        "FIM_REMESSA": "historico ↔ vw_cessao ↔ vw_cessao_fluxo",
        "FIM_ANEXADO": "historico ↔ vw_cessao ↔ vw_cessao_fluxo",
        "RETORNO_UNIDADE": "historico ↔ vw_cessao",
        "REMESSA_SIMULTANEA": "historico ↔ vw_cessao ↔ vw_cessao_fluxo",
        "EPISODIO_EM_ANDAMENTO": "historico ↔ vw_cessao",
    }
    return mapping[class_code]


def flow_evidence_for_occurrence(
    class_code: str,
    proc: str,
    dt: Any,
    hist: pd.DataFrame,
    fluxo: pd.DataFrame,
) -> str:
    if pd.isna(dt):
        return ""

    dt_key = timestamp_key(dt)
    f = fluxo[
        (fluxo["Processo"] == proc)
        & (fluxo["Data_Transicao"].map(timestamp_key) == dt_key)
    ]

    if class_code in {"FIM_CONCLUSAO", "FIM_ANEXADO"}:
        f = f[f["_Destino_Norm"] == "fim"]
    elif class_code in {"FIM_REMESSA", "REMESSA_SIMULTANEA"}:
        f = f[
            (f["_Origem_Norm"] != "início")
            & (f["_Destino_Norm"] != "fim")
        ]
    else:
        return "Não se aplica diretamente à vw_cessao_fluxo."

    if f.empty:
        return "Nenhuma transição correspondente localizada automaticamente."

    parts = []
    for origin, dest, trans_dt in f[
        ["Unidade_Origem", "Unidade_Destino", "Data_Transicao"]
    ].itertuples(index=False, name=None):
        parts.append(
            f"{origin} -> {dest} @ {pd.Timestamp(trans_dt).strftime('%d/%m/%Y %H:%M')}"
        )

    return "; ".join(parts)


def build_manual_behavior_review(
    occurrences: pd.DataFrame,
    selected_processes: list[str],
    hist: pd.DataFrame,
    fluxo: pd.DataFrame,
) -> pd.DataFrame:
    """
    Uma linha por classe observada. Para cada classe, escolhe uma ocorrência
    representativa dentro do conjunto mínimo de processos selecionados.
    """
    selected_occ = occurrences[
        occurrences["Processo"].isin(selected_processes)
    ].copy()

    rows = []

    instructions = {
        "INICIO_GERADO": (
            "Conferir no historico o registro de geração e confirmar que a "
            "vw_cessao contém Data_Inicio compatível para a unidade e instante."
        ),
        "INICIO_RECEBIDO": (
            "Conferir o recebimento no historico e a Data_Inicio correspondente "
            "na vw_cessao."
        ),
        "INICIO_REABERTURA": (
            "Conferir a reabertura no historico e verificar se o marco é "
            "representado adequadamente na passagem correspondente."
        ),
        "FIM_CONCLUSAO": (
            "Conferir a conclusão no historico, a Data_Fim correspondente na "
            "vw_cessao e a transição Unidade -> FIM na vw_cessao_fluxo."
        ),
        "FIM_REMESSA": (
            "Conferir a remessa no historico, a unidade remetente extraída de "
            "Obs, a Data_Fim da vw_cessao e a relação origem -> destino da "
            "vw_cessao_fluxo no mesmo instante."
        ),
        "FIM_ANEXADO": (
            "Conferir a anexação no historico, a Data_Fim correspondente na "
            "vw_cessao e a transição Unidade -> FIM na vw_cessao_fluxo."
        ),
        "RETORNO_UNIDADE": (
            "Conferir no historico que o processo retornou à mesma unidade em "
            "momentos distintos e verificar se a vw_cessao mantém passagens "
            "separadas por StayID."
        ),
        "REMESSA_SIMULTANEA": (
            "Conferir as remessas no mesmo instante e verificar se a "
            "vw_cessao_fluxo preserva todas como transições distintas com "
            "origens e destinos coerentes; verificar também as passagens "
            "correspondentes na vw_cessao quando aplicável."
        ),
        "EPISODIO_EM_ANDAMENTO": (
            "Conferir que não há marco posterior de encerramento aplicável à "
            "passagem e que a vw_cessao mantém Data_Fim vazia e status "
            "Em Andamento."
        ),
    }

    for class_code, class_name in BEHAVIOR_CLASSES:
        subset = selected_occ[
            selected_occ["Classe"] == class_code
        ].sort_values(
            ["Processo", "Data"],
            na_position="last",
        )

        if subset.empty:
            continue

        row = subset.iloc[0]

        rows.append(
            {
                "Classe": class_code,
                "Classe_Descricao": class_name,
                "Processo": row["Processo"],
                "Unidade_Referencia": row["Unidade"],
                "Data_Referencia": row["Data"],
                "Evidencia_historico": row["Evidencia"],
                "Evidencia_vw_fluxo": flow_evidence_for_occurrence(
                    class_code,
                    row["Processo"],
                    row["Data"],
                    hist,
                    fluxo,
                ),
                "Estruturas_Conferidas": describe_structures_for_class(class_code),
                "O_Que_Conferir": instructions[class_code],
                "Resultado_Manual": "",
                "Observacao": "",
            }
        )

    return pd.DataFrame(rows)


# =============================================================================
# PRESERVAÇÃO DA CONFERÊNCIA MANUAL AO SOBRESCREVER A SAÍDA
# =============================================================================


def load_previous_manual_answers() -> dict[tuple[str, str], tuple[str, str]]:
    if not OUTPUT_FILE.exists():
        return {}

    for sheet in ["06_CONFERENCIA_CLASSES", "05_CONFERENCIA_CLASSES"]:
        try:
            previous = pd.read_excel(OUTPUT_FILE, sheet_name=sheet)
        except Exception:
            continue

        required = {"Classe", "Processo", "Resultado_Manual", "Observacao"}
        if not required.issubset(previous.columns):
            continue

        answers: dict[tuple[str, str], tuple[str, str]] = {}
        for cls, proc, result, obs in previous[
            ["Classe", "Processo", "Resultado_Manual", "Observacao"]
        ].itertuples(index=False, name=None):
            key = (str(cls).strip(), normalize_process(proc))
            answers[key] = (
                "" if pd.isna(result) else str(result).strip(),
                "" if pd.isna(obs) else str(obs).strip(),
            )
        return answers

    return {}


def apply_previous_manual_answers(
    manual_review: pd.DataFrame,
    answers: dict[tuple[str, str], tuple[str, str]],
) -> pd.DataFrame:
    out = manual_review.copy()

    if not answers or out.empty:
        return out

    for idx, row in out.iterrows():
        key = (str(row["Classe"]).strip(), normalize_process(row["Processo"]))
        if key in answers:
            result, obs = answers[key]
            out.at[idx, "Resultado_Manual"] = result
            out.at[idx, "Observacao"] = obs

    return out


# =============================================================================
# REGRAS DOCUMENTADAS
# =============================================================================


def rules_dataframe() -> pd.DataFrame:
    rows = [
        (
            "A01",
            "SEI -> historico",
            "Completude da aquisição",
            "Para cada processo, confrontar quantidade de registros, primeiro registro e último registro entre o Histórico do SEI e a tabela historico.",
        ),
        (
            "V01",
            "historico -> vw_cessao",
            "Cobertura do universo",
            "O conjunto de processos de cessão em historico deve coincidir com a vw_cessao.",
        ),
        (
            "V02",
            "historico -> vw_cessao",
            "Rastreabilidade do início",
            "Cada Data_Inicio da vw_cessao deve possuir marco elegível de INÍCIO no historico para o mesmo processo, unidade e instante.",
        ),
        (
            "V03",
            "historico -> vw_cessao",
            "Rastreabilidade do fim",
            "Cada Data_Fim da vw_cessao deve possuir marco elegível de FIM no historico para o mesmo processo, unidade e instante.",
        ),
        (
            "V04",
            "historico -> vw_cessao",
            "Tratamento das remessas",
            "A unidade remetente deve ser identificável em Obs; a correspondência entre essa origem e a unidade do episódio é verificada conjuntamente pela rastreabilidade de Data_Fim.",
        ),
        (
            "V05",
            "historico -> vw_cessao",
            "Coerência temporal e de status",
            "Data_Fim não pode anteceder Data_Inicio; status concluído exige fim e status em andamento exige fim ausente.",
        ),
        (
            "V06",
            "historico -> vw_cessao",
            "Coerência da duração",
            "A duração dos episódios concluídos deve corresponder à diferença cronológica entre Data_Fim e Data_Inicio.",
        ),
        (
            "V07",
            "historico -> vw_cessao",
            "Integridade dos episódios",
            "Campos essenciais não podem estar ausentes e duplicações exatas de episódios são sinalizadas.",
        ),
        (
            "F01",
            "historico -> vw_cessao_fluxo",
            "Cobertura do universo",
            "O conjunto de processos de cessão em historico deve coincidir com a vw_cessao_fluxo.",
        ),
        (
            "F02",
            "historico -> vw_cessao_fluxo",
            "Transição inicial",
            "Cada processo deve possuir uma transição INÍCIO -> unidade inicial no primeiro instante registrado.",
        ),
        (
            "F03",
            "historico -> vw_cessao_fluxo",
            "Remessas origem-destino",
            "As remessas devem preservar, com multiplicidade, a origem extraída de Obs, o destino registrado em Unidade e o timestamp.",
        ),
        (
            "F04",
            "historico -> vw_cessao_fluxo",
            "Encerramentos para FIM",
            "Conclusões e anexações devem produzir relações unidade -> FIM no mesmo timestamp do registro de origem.",
        ),
        (
            "F05",
            "historico -> vw_cessao_fluxo",
            "Origem identificável",
            "A ocorrência de origem DESCONHECIDA é sinalizada para conferência.",
        ),
        (
            "F06",
            "vw_cessao_fluxo",
            "Horas entre transições",
            "Horas_de_Transicao deve ser coerente com a diferença entre timestamps sucessivos; empates são avaliados como multiconjunto no grupo temporal.",
        ),
        (
            "F07",
            "vw_cessao_fluxo",
            "Integridade estrutural",
            "Processo, origem, destino e data devem estar preenchidos e o intervalo não pode ser negativo.",
        ),
        (
            "C01",
            "Controle manual complementar",
            "Cobertura por classes",
            "Conferir ao menos uma ocorrência de cada classe de comportamento efetivamente observada, usando o menor conjunto de processos capaz de cobri-las.",
        ),
    ]

    return pd.DataFrame(
        rows,
        columns=["Codigo", "Etapa", "Controle", "Criterio"],
    )


# =============================================================================
# RESUMO GERAL
# =============================================================================


def build_general_summary(
    hist: pd.DataFrame,
    vw: pd.DataFrame,
    fluxo: pd.DataFrame,
    completeness: pd.DataFrame,
    auto_cessao: pd.DataFrame,
    auto_fluxo: pd.DataFrame,
    class_summary: pd.DataFrame,
    selected_processes: list[str],
    manual_review: pd.DataFrame,
) -> pd.DataFrame:

    hist_processes = hist["Processo"].nunique()
    view_processes = vw["Processo"].nunique()
    flow_processes = fluxo["Processo"].nunique()

    pending = int(
        (completeness["Resultado_Completude"] == "PENDENTE").sum()
    )
    completeness_fail = int(
        (completeness["Resultado_Completude"] == "CONFERIR").sum()
    )
    auto_cessao_fail = int(
        (auto_cessao["Resultado_Automatico"] == "CONFERIR").sum()
    )
    auto_fluxo_fail = int(
        (auto_fluxo["Resultado_Automatico"] == "CONFERIR").sum()
    )

    observed_classes = int(
        (class_summary["Status"] == "OBSERVADA").sum()
    )
    non_observed = int(
        (class_summary["Status"] != "OBSERVADA").sum()
    )

    manual_results = manual_review["Resultado_Manual"].map(normalize_text)
    manual_pending = int((manual_results == "").sum())
    manual_conferir = int((manual_results == "conferir").sum())
    manual_ok = int((manual_results == "ok").sum())

    if pending > 0:
        overall = "PENDENTE COMPLETUDE MANUAL"
    elif completeness_fail > 0 or auto_cessao_fail > 0 or auto_fluxo_fail > 0:
        overall = "CONFERIR DIVERGÊNCIAS AUTOMÁTICAS"
    elif manual_conferir > 0:
        overall = "CONFERIR CONTROLE MANUAL"
    elif manual_pending > 0:
        overall = "PENDENTE CONTROLE MANUAL POR CLASSES"
    else:
        overall = "OK"

    values = [
        ("Data/hora da execução", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ("Processos esperados", EXPECTED_PROCESS_COUNT),
        ("Processos de cessão em historico", hist_processes),
        ("Processos na vw_cessao", view_processes),
        ("Processos na vw_cessao_fluxo", flow_processes),
        (
            "historico e vw_cessao possuem o mesmo conjunto de processos",
            "SIM" if set(hist["Processo"].unique()) == set(vw["Processo"].unique()) else "NÃO",
        ),
        (
            "historico e vw_cessao_fluxo possuem o mesmo conjunto de processos",
            "SIM" if set(hist["Processo"].unique()) == set(fluxo["Processo"].unique()) else "NÃO",
        ),
        ("Registros de historico analisados", len(hist)),
        ("Episódios da vw_cessao analisados", len(vw)),
        ("Transições da vw_cessao_fluxo analisadas", len(fluxo)),
        ("Processos com completude pendente", pending),
        ("Processos com divergência de completude", completeness_fail),
        ("Processos com falha automática na vw_cessao", auto_cessao_fail),
        ("Processos com falha automática na vw_cessao_fluxo", auto_fluxo_fail),
        ("Classes de comportamento previstas", len(BEHAVIOR_CLASSES)),
        ("Classes observadas no recorte", observed_classes),
        ("Classes não observadas no recorte", non_observed),
        (
            "Processos necessários para cobrir todas as classes observadas",
            len(selected_processes),
        ),
        ("Processos selecionados", ", ".join(selected_processes)),
        ("Ocorrências previstas para conferência manual", len(manual_review)),
        ("Conferências manuais OK", manual_ok),
        ("Conferências manuais pendentes", manual_pending),
        ("Conferências manuais marcadas CONFERIR", manual_conferir),
        ("Fuso horário de normalização", LOCAL_TIMEZONE),
        ("Resultado geral", overall),
    ]

    return pd.DataFrame(
        values,
        columns=["Indicador", "Valor"],
    )


# =============================================================================
# EXCEL
# =============================================================================


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)

    header_fill = PatternFill(
        "solid",
        fgColor="D9E1F2",
    )

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        for column_cells in ws.columns:
            max_len = 0
            col_letter = column_cells[0].column_letter

            for cell in column_cells[:200]:
                if cell.value is not None:
                    max_len = max(
                        max_len,
                        len(str(cell.value)),
                    )

            ws.column_dimensions[col_letter].width = min(
                max(max_len + 2, 12),
                55,
            )

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

    # Lista padronizada no campo Resultado_Manual.
    if "06_CONFERENCIA_CLASSES" in wb.sheetnames:
        ws = wb["06_CONFERENCIA_CLASSES"]
        headers = {cell.value: cell.column for cell in ws[1]}
        col = headers.get("Resultado_Manual")
        if col and ws.max_row >= 2:
            dv = DataValidation(
                type="list",
                formula1='"OK,CONFERIR"',
                allow_blank=True,
            )
            ws.add_data_validation(dv)
            dv.add(f"{ws.cell(2, col).coordinate}:{ws.cell(ws.max_row, col).coordinate}")

    wb.save(path)


def export_report(
    general: pd.DataFrame,
    completeness: pd.DataFrame,
    auto_cessao: pd.DataFrame,
    auto_fluxo: pd.DataFrame,
    divergences: pd.DataFrame,
    class_summary: pd.DataFrame,
    manual_review: pd.DataFrame,
    rules: pd.DataFrame,
) -> None:

    frames = [
        make_excel_safe(df)
        for df in [
            general,
            completeness,
            auto_cessao,
            auto_fluxo,
            divergences,
            class_summary,
            manual_review,
            rules,
        ]
    ]

    (
        general,
        completeness,
        auto_cessao,
        auto_fluxo,
        divergences,
        class_summary,
        manual_review,
        rules,
    ) = frames

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl",
        datetime_format="dd/mm/yyyy hh:mm:ss",
        date_format="dd/mm/yyyy",
    ) as writer:

        general.to_excel(
            writer,
            index=False,
            sheet_name="00_RESUMO_GERAL",
        )

        completeness.to_excel(
            writer,
            index=False,
            sheet_name="01_COMPLETUDE_EXTRACAO",
        )

        auto_cessao.to_excel(
            writer,
            index=False,
            sheet_name="02_VALIDACAO_VW_CESSAO",
        )

        auto_fluxo.to_excel(
            writer,
            index=False,
            sheet_name="03_VALIDACAO_VW_FLUXO",
        )

        divergences.to_excel(
            writer,
            index=False,
            sheet_name="04_DIVERGENCIAS",
        )

        class_summary.to_excel(
            writer,
            index=False,
            sheet_name="05_CLASSES_COMPORTAMENTO",
        )

        manual_review.to_excel(
            writer,
            index=False,
            sheet_name="06_CONFERENCIA_CLASSES",
        )

        rules.to_excel(
            writer,
            index=False,
            sheet_name="07_CONTROLES",
        )

    style_workbook(OUTPUT_FILE)


# =============================================================================
# EXECUÇÃO
# =============================================================================


def main():
    print("Validação V7")
    print(f"Pasta do script: {BASE_DIR}")
    print(f"Fuso horário: {LOCAL_TIMEZONE}")

    print("Conectando ao PostgreSQL...")
    engine = get_engine()

    print("Lendo historico, vw_cessao e vw_cessao_fluxo...")
    hist, vw, fluxo = load_data(engine)

    print("Preparando dados...")
    hist, vw, fluxo = prepare_data(hist, vw, fluxo)

    hist_processes = set(hist["Processo"].dropna().unique())
    view_processes = set(vw["Processo"].dropna().unique())
    flow_processes = set(fluxo["Processo"].dropna().unique())
    processes = sorted(hist_processes | view_processes | flow_processes)

    print(f"Processos em historico: {len(hist_processes)}")
    print(f"Processos em vw_cessao: {len(view_processes)}")
    print(f"Processos em vw_cessao_fluxo: {len(flow_processes)}")

    created = ensure_manual_input(processes)

    if created:
        print(f"Arquivo de conferência criado: {MANUAL_INPUT}")
        print(
            "Preencha Qtd_Registros_SEI, Primeiro_Registro_SEI e "
            "Ultimo_Registro_SEI e execute novamente."
        )

    print("1/4 - Montando controle de completude...")
    completeness = build_completeness(hist, processes)

    print("2/4 - Validando vw_cessao contra historico...")
    auto_cessao, div_cessao = run_automatic_validation_cessao(hist, vw)

    print("3/4 - Validando vw_cessao_fluxo contra historico...")
    auto_fluxo, div_fluxo = run_automatic_validation_fluxo(hist, fluxo)

    divergences = pd.concat(
        [div_cessao, div_fluxo],
        ignore_index=True,
    )

    print("4/4 - Identificando classes de comportamento...")
    occurrences = build_behavior_occurrences(hist, vw)
    class_summary = build_class_summary(occurrences)

    selected_processes = select_minimum_process_cover(occurrences)

    manual_review = build_manual_behavior_review(
        occurrences,
        selected_processes,
        hist,
        fluxo,
    )

    previous_answers = load_previous_manual_answers()
    manual_review = apply_previous_manual_answers(
        manual_review,
        previous_answers,
    )

    print(
        "Processos selecionados para cobrir todas as classes observadas: "
        f"{len(selected_processes)}"
    )

    for proc in selected_processes:
        print(f"  - {proc}")

    general = build_general_summary(
        hist,
        vw,
        fluxo,
        completeness,
        auto_cessao,
        auto_fluxo,
        class_summary,
        selected_processes,
        manual_review,
    )

    print("Gerando planilha consolidada...")
    print(f"A saída será sobrescrita em: {OUTPUT_FILE}")

    export_report(
        general,
        completeness,
        auto_cessao,
        auto_fluxo,
        divergences,
        class_summary,
        manual_review,
        rules_dataframe(),
    )

    print(f"Planilha gerada: {OUTPUT_FILE}")
    print("Concluído.")


if __name__ == "__main__":
    main()
